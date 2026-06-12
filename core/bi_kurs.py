"""
SAKSI — Penarik kurs Bank Indonesia (web service ``wskursbi``)
==============================================================
Menarik **Kurs Transaksi BI** (``getSubKursLokal3``) dan/atau **JISDOR**
(``getSubKursJisdor3``) untuk satu rentang tanggal & daftar mata uang pilihan,
lalu merapikannya menjadi DataFrame yang siap diekspor ke Excel.

Catatan teknis: WAF Bank Indonesia me-reset fingerprint TLS non-browser
(``requests``/``urllib`` kena reset saat handshake), sedangkan handshake ``curl``
diterima. Karena itu pengambilan dilakukan lewat ``curl`` (tersedia bawaan di
Windows 10+, macOS, dan Linux).

Sumber contoh:
https://www.bi.go.id/biwebservice/wskursbi.asmx/getSubKursJisdor3?mts=USD&startDate=2026-01-01&endDate=2026-06-09
"""
from __future__ import annotations

import io
import subprocess
import xml.etree.ElementTree as ET
from typing import Callable, Iterable

import pandas as pd

# Daftar mata uang yang disediakan BI di web service kurs.
CURRENCIES = [
    "AED", "AUD", "BND", "CAD", "CHF", "CNH", "CNY", "DKK", "EUR", "GBP",
    "HKD", "JPY", "KRW", "KWD", "LAK", "MYR", "NOK", "NZD", "PGK", "PHP",
    "SAR", "SEK", "SGD", "THB", "USD", "VND",
]

BASE_URL = "https://www.bi.go.id/biwebservice/wskursbi.asmx"
M_TRANSAKSI = "getSubKursLokal3"   # Kurs Transaksi BI (beli/jual per valuta)
M_JISDOR = "getSubKursJisdor3"     # JISDOR (referensi USD/IDR & valuta lain)

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")


# ----------------------------------------------------------------------------
# Pengambilan & parsing mentah
# ----------------------------------------------------------------------------
def fetch(method: str, code: str, start: str, end: str) -> str:
    """Ambil XML mentah satu valuta lewat curl. ``start``/``end`` = 'YYYY-MM-DD'."""
    url = f"{BASE_URL}/{method}?mts={code}&startdate={start}&enddate={end}"
    result = subprocess.run(
        ["curl", "-sS", "--max-time", "60", "-A", UA, url],
        capture_output=True, text=True, encoding="utf-8",
    )
    if result.returncode != 0:
        raise RuntimeError(f"curl gagal untuk {method}/{code}: {result.stderr.strip()}")
    return result.stdout


def parse_bi_xml(xml_text: str, code: str) -> pd.DataFrame:
    """Generik: ambil semua baris ``<Table>`` apa pun suffix kolomnya."""
    rows = []
    root = ET.fromstring(xml_text)
    for elem in root.iter():
        if elem.tag.split("}")[-1] != "Table":
            continue
        row = {"Kode": code}
        for c in elem:
            tag = c.tag.split("}")[-1]
            row[tag] = c.text
        rows.append(row)
    return pd.DataFrame(rows)


# ----------------------------------------------------------------------------
# Perapian → pertahankan SEMUA kolom mentah BI + tambah Kurs Tengah & Fix Date
# Susunan kolom akhir (sesuai tampilan Excel acuan):
#   Kode, id_*, lnk_*, nil_*, beli_*, jual_*, tgl_*, mts_*, Kurs Tengah, Fix Date
# ----------------------------------------------------------------------------
def _first(cols: Iterable[str], prefix: str) -> str | None:
    return next((c for c in cols if c.startswith(prefix)), None)


def _tidy(df: pd.DataFrame) -> pd.DataFrame:
    """Pertahankan kolom mentah; angka jadi numerik; tambah Kurs Tengah & Fix Date."""
    if df.empty:
        return df
    df = df.copy()
    beli, jual = _first(df.columns, "beli"), _first(df.columns, "jual")
    nil, tgl = _first(df.columns, "nil"), _first(df.columns, "tgl")
    for c in (nil, beli, jual):
        if c:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    if beli and jual:
        df["Kurs Tengah"] = (df[beli] + df[jual]) / 2
    if tgl:
        df["Fix Date"] = pd.to_datetime(df[tgl].str.slice(0, 10), errors="coerce")
    return df


# ----------------------------------------------------------------------------
# Orkestrasi multi-valuta
# ----------------------------------------------------------------------------
def _tarik(method: str, codes: list[str], start: str, end: str,
           progress: Callable[[int, int, str], None] | None) -> pd.DataFrame:
    parts = []
    total = len(codes)
    for i, code in enumerate(codes, 1):
        if progress:
            progress(i, total, code)
        df = _tidy(parse_bi_xml(fetch(method, code, start, end), code))
        if not df.empty:
            parts.append(df)
    if not parts:
        return pd.DataFrame()
    out = pd.concat(parts, ignore_index=True)
    sort_cols = ["Kode"] + (["Fix Date"] if "Fix Date" in out.columns else [])
    return out.sort_values(sort_cols).reset_index(drop=True)


def tarik_kurs_transaksi(codes, start, end, progress=None) -> pd.DataFrame:
    """Kurs Transaksi BI untuk daftar valuta & rentang tanggal."""
    return _tarik(M_TRANSAKSI, list(codes), start, end, progress)


def tarik_jisdor(codes, start, end, progress=None) -> pd.DataFrame:
    """JISDOR untuk daftar valuta & rentang tanggal."""
    return _tarik(M_JISDOR, list(codes), start, end, progress)


# ----------------------------------------------------------------------------
# Ekspor Excel
# ----------------------------------------------------------------------------
_RP_FMT = '"Rp"#,##0.00'
_DATE_FMT = "m/d/yyyy"


def build_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Susun workbook .xlsx dari {nama_sheet: DataFrame}. Sheet kosong dilewati.

    Kolom nilai (beli/jual/Kurs Tengah) diberi format mata uang Rp; 'Fix Date'
    diberi format tanggal — meniru tampilan Excel acuan."""
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        wrote = False
        for name, df in sheets.items():
            if df is None or df.empty:
                continue
            sn = name[:31]
            df.to_excel(writer, sheet_name=sn, index=False)
            ws = writer.sheets[sn]
            for ci, col in enumerate(df.columns, start=1):
                if col.startswith(("beli", "jual")) or col == "Kurs Tengah":
                    fmt = _RP_FMT
                elif col == "Fix Date":
                    fmt = _DATE_FMT
                else:
                    continue
                letter = get_column_letter(ci)
                for cell in ws[letter][1:]:  # lewati baris header
                    cell.number_format = fmt
            wrote = True
        if not wrote:
            pd.DataFrame({"info": ["Tidak ada data pada rentang/valuta terpilih"]}).to_excel(
                writer, sheet_name="Kosong", index=False)
    buf.seek(0)
    return buf.getvalue()
